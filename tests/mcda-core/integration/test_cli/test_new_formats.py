"""
CLI 新格式集成测试

测试 CLI 对新报告格式的支持：
- HTML 报告
- PDF 报告
- Excel 导出
"""

import sys
import tempfile
from pathlib import Path

import pytest

# 添加 mcda_core 模块路径
mcda_core_path = Path(__file__).parent.parent.parent.parent / "skills" / "mcda-core" / "lib"
if str(mcda_core_path) not in sys.path:
    sys.path.insert(0, str(mcda_core_path))

from mcda_core.cli import MCDACommandLineInterface


# ============================================================================
# Test Fixtures
# ============================================================================

@pytest.fixture
def sample_config_file():
    """创建示例配置文件"""
    config_content = """# 示例供应商选择配置

name: 供应商选择决策

alternatives:
  - 供应商A
  - 供应商B
  - 供应商C

criteria:
  - name: 成本
    weight: 0.4
    direction: lower_better
    description: 采购成本（越低越好）

  - name: 质量
    weight: 0.3
    direction: higher_better
    description: 产品质量评分（越高越好）

  - name: 交货期
    weight: 0.3
    direction: lower_better
    description: 交付时间（天数，越短越好）

scores:
  供应商A:
    成本: 50
    质量: 80
    交货期: 30

  供应商B:
    成本: 60
    质量: 70
    交货期: 25

  供应商C:
    成本: 45
    质量: 85
    交货期: 35

algorithm:
  name: topsis
"""

    with tempfile.NamedTemporaryFile(mode="w", suffix=".yaml", delete=False) as f:
        f.write(config_content)
        return f.name


@pytest.fixture
def cli():
    """CLI 实例"""
    return MCDACommandLineInterface()


# ============================================================================
# HTML 格式测试
# ============================================================================

def test_html_format_via_cli(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试通过 CLI 生成 HTML 报告"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "report.html"

        # 运行 CLI 命令
        args = [
            "analyze",
            sample_config_file,
            "-o", str(output_file),
            "-f", "html",
        ]

        cli.run(args)

        # 验证文件存在
        assert output_file.exists()

        # 验证文件内容
        content = output_file.read_text(encoding="utf-8")
        assert "<!DOCTYPE html>" in content
        assert "<html" in content
        assert "供应商选择" in content


def test_html_format_with_chart(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试生成包含图表的 HTML 报告"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "report_with_chart.html"

        # 运行 CLI 命令
        args = [
            "analyze",
            sample_config_file,
            "-o", str(output_file),
            "-f", "html",
            "--include-chart",
        ]

        cli.run(args)

        # 验证文件存在
        assert output_file.exists()

        # 验证包含图表（base64 编码的图片）
        content = output_file.read_text(encoding="utf-8")
        assert "data:image/png;base64" in content


# ============================================================================
# PDF 格式测试
# ============================================================================

def test_pdf_format_via_cli(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试通过 CLI 生成 PDF 报告"""
    try:
        import weasyprint  # noqa: F401
        WEASYPRINT_AVAILABLE = True
    except ImportError:
        WEASYPRINT_AVAILABLE = False

    if not WEASYPRINT_AVAILABLE:
        pytest.skip("weasyprint 未安装")

    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "report.pdf"

        # 运行 CLI 命令
        args = [
            "analyze",
            sample_config_file,
            "-o", str(output_file),
            "-f", "pdf",
        ]

        cli.run(args)

        # 验证文件存在
        assert output_file.exists()

        # 验证是有效的 PDF 文件
        content = output_file.read_bytes()
        assert content[:4] == b"%PDF"


# ============================================================================
# Excel 格式测试
# ============================================================================

def test_excel_format_via_cli(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试通过 CLI 导出 Excel"""
    import openpyxl

    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "report.xlsx"

        # 运行 CLI 命令
        args = [
            "analyze",
            sample_config_file,
            "-o", str(output_file),
            "-f", "excel",
        ]

        cli.run(args)

        # 验证文件存在
        assert output_file.exists()

        # 验证是有效的 Excel 文件
        wb = openpyxl.load_workbook(output_file)
        assert wb is not None

        # 验证工作表
        assert "Overview" in wb.sheetnames
        assert "Rankings" in wb.sheetnames
        assert "Scores Matrix" in wb.sheetnames

        wb.close()


# ============================================================================
# 多格式输出测试
# ============================================================================

def test_multiple_formats(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试同时生成多种格式的报告"""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # 生成 HTML
        html_file = tmpdir / "report.html"
        cli.run(["analyze", sample_config_file, "-o", str(html_file), "-f", "html"])
        assert html_file.exists()

        # 生成 Excel
        excel_file = tmpdir / "report.xlsx"
        cli.run(["analyze", sample_config_file, "-o", str(excel_file), "-f", "excel"])
        assert excel_file.exists()

        # 生成 Markdown（原有格式）
        md_file = tmpdir / "report.md"
        cli.run(["analyze", sample_config_file, "-o", str(md_file), "-f", "markdown"])
        assert md_file.exists()


# ============================================================================
# 向后兼容性测试
# ============================================================================

def test_backward_compatibility_markdown(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试原有 Markdown 格式仍然可用"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "report.md"

        # 运行 CLI 命令
        args = [
            "analyze",
            sample_config_file,
            "-o", str(output_file),
            "-f", "markdown",
        ]

        cli.run(args)

        # 验证文件存在
        assert output_file.exists()

        # 验证 Markdown 内容
        content = output_file.read_text(encoding="utf-8")
        assert "# " in content  # Markdown 标题
        assert "## 决策结果" in content


def test_backward_compatibility_json(
    cli: MCDACommandLineInterface,
    sample_config_file: str,
):
    """测试原有 JSON 格式仍然可用"""
    with tempfile.TemporaryDirectory() as tmpdir:
        output_file = Path(tmpdir) / "report.json"

        # 运行 CLI 命令
        args = [
            "analyze",
            sample_config_file,
            "-o", str(output_file),
            "-f", "json",
        ]

        cli.run(args)

        # 验证文件存在
        assert output_file.exists()

        # 验证 JSON 内容
        import json
        with open(output_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            assert "problem" in data
            assert "result" in data
