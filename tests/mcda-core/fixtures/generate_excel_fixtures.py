"""
生成 Excel 测试数据文件

这个脚本使用 openpyxl 创建测试用的 Excel 文件。
运行前请确保已安装 openpyxl：pip install openpyxl
"""

from pathlib import Path


def create_standard_excel():
    """创建标准 Excel 测试文件"""
    try:
        import openpyxl
    except ImportError:
        print("错误: openpyxl 未安装，请运行: pip install openpyxl")
        return False

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入标题行
    ws.append(['', '权重', '方向', '方案A', '方案B', '方案C'])

    # 写入准则数据
    ws.append(['性能', 0.4, 'higher', 85, 90, 88])
    ws.append(['成本', 0.3, 'lower', 50, 60, 55])
    ws.append(['可靠性', 0.2, 'higher', 90, 85, 92])
    ws.append(['易用性', 0.1, 'higher', 80, 75, 78])

    # 保存
    output_path = Path(__file__).parent / 'decision_data.xlsx'
    wb.save(output_path)
    print(f"✅ 已创建: {output_path}")
    return True


def create_interval_excel():
    """创建区间数 Excel 测试文件"""
    try:
        import openpyxl
    except ImportError:
        print("错误: openpyxl 未安装，请运行: pip install openpyxl")
        return False

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 写入标题行
    ws.append(['', '权重', '方向', '方案A', '方案B', '方案C'])

    # 写入准则数据（区间数）
    ws.append(['性能', 0.4, 'higher', '80,90', '85,95', '82,92'])
    ws.append(['成本', 0.3, 'lower', '45,55', '50,65', '48,58'])
    ws.append(['可靠性', 0.2, 'higher', '88,92', '82,88', '90,94'])
    ws.append(['易用性', 0.1, 'higher', '75,85', '70,80', '75,82'])

    # 保存
    output_path = Path(__file__).parent / 'decision_data_interval.xlsx'
    wb.save(output_path)
    print(f"✅ 已创建: {output_path}")
    return True


def create_template_excel():
    """创建 Excel 模板文件"""
    try:
        import openpyxl
    except ImportError:
        print("错误: openpyxl 未安装，请运行: pip install openpyxl")
        return False

    # 创建工作簿
    wb = openpyxl.Workbook()

    # Sheet1: 决策矩阵模板
    ws = wb.active
    ws.title = "决策矩阵"

    # 写入标题行
    ws.append(['', '权重', '方向', '方案A', '方案B', '方案C'])

    # 写入示例数据
    ws.append(['性能', 0.4, 'higher', 85, 90, 88])
    ws.append(['成本', 0.3, 'lower', 50, 60, 55])
    ws.append(['可靠性', 0.2, 'higher', 90, 85, 92])
    ws.append(['易用性', 0.1, 'higher', 80, 75, 78])

    # Sheet2: 元信息
    ws_meta = wb.create_sheet("元信息")
    ws_meta.append(['项目', '内容'])
    ws_meta.append(['问题名称', '供应商选择'])
    ws_meta.append(['算法', 'topsis'])
    ws_meta.append(['描述', '使用 TOPSIS 算法进行供应商选择'])

    # 保存到模板目录
    template_dir = Path(__file__).parent.parent.parent.parent / 'docs' / 'active' / 'mcda-core' / 'v0.9' / 'templates'
    template_dir.mkdir(parents=True, exist_ok=True)
    output_path = template_dir / 'excel_template.xlsx'
    wb.save(output_path)
    print(f"✅ 已创建: {output_path}")
    return True


if __name__ == '__main__':
    print("正在生成 Excel 测试数据文件...")
    print()

    success = True
    success &= create_standard_excel()
    success &= create_interval_excel()
    success &= create_template_excel()

    if success:
        print()
        print("=" * 50)
        print("✅ 所有 Excel 文件创建成功！")
        print("=" * 50)
    else:
        print()
        print("=" * 50)
        print("❌ 部分文件创建失败，请检查 openpyxl 是否已安装")
        print("=" * 50)
