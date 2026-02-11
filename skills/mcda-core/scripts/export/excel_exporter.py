"""
Excel 导出器

提供 Excel 格式的决策结果导出功能：
- 多工作表导出
- 单元格格式化
- 数据验证
"""

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

if TYPE_CHECKING:
    from ..models import DecisionProblem, DecisionResult


class ExcelExporter:
    """Excel 导出器

    将决策问题和结果导出为格式化的 Excel 文件。
    """

    def __init__(self):
        """初始化 Excel 导出器"""
        # 样式定义
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def export_excel(
        self,
        problem: "DecisionProblem",
        result: "DecisionResult",
        *,
        include_charts: bool = False,
    ) -> bytes:
        """
        导出 Excel 字节流

        Args:
            problem: 决策问题
            result: 决策结果
            include_charts: 是否包含图表（暂不支持）

        Returns:
            bytes: Excel 字节流
        """
        # 创建工作簿
        wb = openpyxl.Workbook()

        # 删除默认工作表
        wb.remove(wb.active)

        # 创建各个工作表
        self._create_overview_sheet(wb, problem, result)
        self._create_rankings_sheet(wb, result)
        self._create_scores_matrix_sheet(wb, problem)

        # 保存到字节流
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()

    def save_excel(
        self,
        problem: "DecisionProblem",
        result: "DecisionResult",
        file_path: str,
        *,
        include_charts: bool = False,
    ) -> None:
        """
        保存 Excel 文件

        Args:
            problem: 决策问题
            result: 决策结果
            file_path: 文件路径
            include_charts: 是否包含图表

        Raises:
            IOError: 文件保存失败
        """
        excel_bytes = self.export_excel(
            problem,
            result,
            include_charts=include_charts,
        )

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with open(path, "wb") as f:
            f.write(excel_bytes)

    def _create_overview_sheet(
        self,
        wb: openpyxl.Workbook,
        problem: "DecisionProblem",
        result: "DecisionResult",
    ) -> None:
        """创建 Overview 工作表"""
        ws = wb.create_sheet("Overview", 0)

        # 标题
        ws["A1"] = "MCDA 决策分析报告"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:B1")

        # 生成时间
        ws["A3"] = "生成时间:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 备选方案数
        ws["A5"] = "备选方案数:"
        ws["B5"] = len(problem.alternatives)

        # 准则数
        ws["A6"] = "准则数:"
        ws["B6"] = len(problem.criteria)

        # 算法信息
        ws["A8"] = "算法信息"
        ws["A8"].font = Font(size=12, bold=True)

        ws["A9"] = "算法名称:"
        ws["B9"] = result.metadata.algorithm_name

        # 备选方案列表
        ws["A11"] = "备选方案"
        ws["A11"].font = Font(size=12, bold=True)

        for i, alt in enumerate(problem.alternatives, start=12):
            ws[f"A{i}"] = alt

        # 调整列宽
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

    def _create_rankings_sheet(
        self,
        wb: openpyxl.Workbook,
        result: "DecisionResult",
    ) -> None:
        """创建 Rankings 工作表"""
        ws = wb.create_sheet("Rankings", 1)

        # 表头
        headers = ["排名", "方案", "评分"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 数据行
        for ranking in result.rankings:
            row = ranking.rank + 1
            ws.cell(row=row, column=1, value=ranking.rank)
            ws.cell(row=row, column=2, value=ranking.alternative)
            ws.cell(row=row, column=3, value=round(ranking.score, 4))

            # 添加边框
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.border

        # 调整列宽
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_scores_matrix_sheet(
        self,
        wb: openpyxl.Workbook,
        problem: "DecisionProblem",
    ) -> None:
        """创建 Scores Matrix 工作表"""
        ws = wb.create_sheet("Scores Matrix", 2)

        # 表头（第一行：方案，第二行：准则）
        ws.cell(row=1, column=1, value="方案 \\ 准则")

        # 第一行：准则名称
        for col, crit in enumerate(problem.criteria, start=2):
            cell = ws.cell(row=1, column=col, value=crit.name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 数据行
        for row, alt in enumerate(problem.alternatives, start=2):
            # 第一列：方案名称
            cell = ws.cell(row=row, column=1, value=alt)
            cell.font = Font(bold=True)
            cell.border = self.border

            # 后续列：评分
            for col, crit in enumerate(problem.criteria, start=2):
                score = problem.scores[alt][crit.name]
                cell = ws.cell(row=row, column=col, value=round(score, 1))
                cell.border = self.border

        # 调整列宽
        ws.column_dimensions["A"].width = 15
        for col in range(2, len(problem.criteria) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
