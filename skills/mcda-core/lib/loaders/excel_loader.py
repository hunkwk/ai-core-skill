"""
Excel 配置文件加载器

支持从 Excel 文件加载决策数据，包括区间数格式。

Excel 格式约定：

Sheet1: 决策矩阵
|         | 权重   | 方向   | 方案A | 方案B | 方案C |
|---------|--------|--------|-------|-------|-------|
| 性能    | 0.4    | higher | 85    | 90    | 88    |
| 成本    | 0.3    | lower  | 50    | 60    | 55    |
| 可靠性  | 0.2    | higher | 90    | 85    | 92    |
| 易用性  | 0.1    | higher | 80    | 75    | 78    |

Sheet2: 元信息（可选）
| 项目     | 内容       |
|---------|-----------|
| 问题名称 | 供应商选择 |
| 算法     | topsis    |
| 描述     | ...       |

区间数支持两种格式：
- Excel 单元格：80,90 或 [80,90]
- 跨单元格：左上角=80，右上角=90（使用命名区域）
"""

from pathlib import Path
from typing import Any, Union

from . import ConfigLoader


class ExcelLoader(ConfigLoader):
    """Excel 配置文件加载器"""

    # 方向别名映射（与 CSV 一致）
    DIRECTION_ALIASES = {
        'higher': 'higher',
        'h': 'higher',
        'higher_is_better': 'higher',
        'lower': 'lower',
        'l': 'lower',
        'lower_is_better': 'lower',
    }

    def __init__(self):
        """初始化 Excel Loader"""
        super().__init__()
        self.alternatives: list[str] = []
        self.criteria: list[dict[str, Any]] = []
        self.matrix: list[list[Any]] = []
        self.metadata: dict[str, Any] = {}

    def load(
        self,
        source: Union[str, Path],
        sheet: Union[str, int] = 0
    ) -> dict[str, Any]:
        """
        加载 Excel 配置文件

        Args:
            source: Excel 文件路径（.xlsx 格式）
            sheet: Sheet 名称或索引，默认为 0（第一个 Sheet）

        Returns:
            解析后的配置字典

        Raises:
            FileNotFoundError: 文件不存在
            ImportError: openpyxl 未安装
            ValueError: 文件格式错误
        """
        # 检查 openpyxl 是否安装
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError(
                "Excel 文件处理需要 openpyxl 库，请安装："
                "pip install openpyxl"
            ) from e

        source_path = Path(source)
        if not source_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {source}")

        if source_path.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(
                f"不支持的文件格式: {source_path.suffix}，"
                f"仅支持 .xlsx 和 .xlsm 格式"
            )

        # 加载工作簿
        try:
            wb = openpyxl.load_workbook(source_path, data_only=True)
        except Exception as e:
            raise ValueError(
                f"无法读取 Excel 文件: {source}"
            ) from e

        # 获取目标 Sheet
        try:
            ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
        except (KeyError, IndexError) as e:
            available_sheets = [ws.title for ws in wb.worksheets]
            raise ValueError(
                f"Sheet '{sheet}' 不存在，"
                f"可用的 Sheet: {available_sheets}"
            ) from e

        # 解析决策矩阵
        self._parse_decision_matrix(ws)

        # 尝试读取元信息 Sheet（可选）
        self._parse_metadata_sheet(wb)

        # 构建配置字典
        config = {
            'alternatives': self.alternatives,
            'criteria': self.criteria,
            'matrix': self.matrix,
            'metadata': {
                'source': str(source_path),
                'format': 'excel',
                'sheet': ws.title,
                **self.metadata,
            }
        }

        wb.close()
        return config

    def _parse_decision_matrix(self, worksheet) -> None:
        """
        解析决策矩阵 Sheet

        Args:
            worksheet: openpyxl Worksheet 对象

        Raises:
            ValueError: 数据格式错误
        """
        # 获取所有数据
        data = []
        for row in worksheet.iter_rows(values_only=True):
            # 过滤空行
            if any(cell is not None for cell in row):
                data.append(row)

        # 验证基本格式
        if len(data) < 2:
            raise ValueError(
                f"Sheet '{worksheet.title}' 格式错误："
                f"至少需要 2 行数据（标题行 + 至少 1 个准则行）"
            )

        # 第一行：标题行（空, 权重, 方向, 方案A, 方案B, ...）
        header = data[0]
        if len(header) < 4:
            raise ValueError(
                f"Sheet '{worksheet.title}' 格式错误："
                f"标题行至少需要 4 列（权重, 方向, 至少 1 个方案）"
            )

        # 提取备选方案名称（从第 4 列开始）
        self.alternatives = [str(cell) for cell in header[3:] if cell is not None]
        if not self.alternatives:
            raise ValueError(
                f"Sheet '{worksheet.title}' 格式错误：未找到备选方案"
            )

        # 解析准则数据（从第 2 行开始）
        self.criteria = []
        self.matrix = []

        for row_idx, row in enumerate(data[1:], start=2):
            if len(row) < 4:
                raise ValueError(
                    f"Sheet '{worksheet.title}' 第 {row_idx} 行格式错误："
                    f"至少需要 4 列（准则名称, 权重, 方向, 得分）"
                )

            criterion_name = str(row[0]) if row[0] else f"准则{row_idx-1}"
            weight = row[1]
            direction = row[2]
            scores = row[3:3+len(self.alternatives)]

            # 验证权重
            if weight is None:
                raise ValueError(
                    f"Sheet '{worksheet.title}' 第 {row_idx} 行错误：权重值不能为空"
                )
            try:
                weight = float(weight)
                if weight < 0:
                    raise ValueError(f"权重不能为负数: {weight}")
            except (TypeError, ValueError) as e:
                raise ValueError(
                    f"Sheet '{worksheet.title}' 第 {row_idx} 行错误："
                    f"权重值无效 '{weight}'"
                ) from e

            # 解析方向
            if direction is None:
                raise ValueError(
                    f"Sheet '{worksheet.title}' 第 {row_idx} 行错误：方向值不能为空"
                )
            direction_str = str(direction).lower()
            if direction_str not in self.DIRECTION_ALIASES:
                raise ValueError(
                    f"Sheet '{worksheet.title}' 第 {row_idx} 行错误："
                    f"方向值无效 '{direction}'，支持的值：higher, lower (或 h, l)"
                )
            direction = self.DIRECTION_ALIASES[direction_str]

            # 验证得分数量
            if len(scores) != len(self.alternatives):
                raise ValueError(
                    f"Sheet '{worksheet.title}' 第 {row_idx} 行错误："
                    f"得分数量与备选方案数量不匹配，"
                    f"期望 {len(self.alternatives)} 个，实际 {len(scores)} 个"
                )

            # 解析得分（支持区间数）
            parsed_scores = []
            for alt_idx, score in enumerate(scores, start=1):
                if score is None:
                    raise ValueError(
                        f"Sheet '{worksheet.title}' 第 {row_idx} 行第 {alt_idx+3} 列错误："
                        f"得分值不能为空"
                    )
                try:
                    parsed_score = self._parse_score(score, row_idx, alt_idx + 3)
                    parsed_scores.append(parsed_score)
                except ValueError as e:
                    raise ValueError(
                        f"Sheet '{worksheet.title}' 第 {row_idx} 行第 {alt_idx+3} 列错误："
                        f"得分值无效 '{score}'"
                    ) from e

            # 添加到结果
            self.criteria.append({
                'name': criterion_name,
                'weight': weight,
                'direction': direction,
            })
            self.matrix.append(parsed_scores)

    def _parse_metadata_sheet(self, workbook) -> None:
        """
        解析元信息 Sheet（可选）

        Args:
            workbook: openpyxl Workbook 对象
        """
        # 尝试查找名为 "元信息" 或 "metadata" 的 Sheet
        metadata_sheet_names = ['元信息', 'metadata', 'meta', 'info']

        for sheet_name in metadata_sheet_names:
            if sheet_name in workbook.sheetnames:
                try:
                    ws = workbook[sheet_name]
                    self.metadata = {}
                    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                        if row[0] and row[1]:
                            key = str(row[0]).strip()
                            value = str(row[1]).strip()
                            self.metadata[key] = value
                    break
                except Exception:
                    # 如果读取失败，忽略元信息
                    pass

    def _parse_score(self, score_value: Any, row_idx: int, col_idx: int) -> Any:
        """
        解析得分值（支持区间数）

        Args:
            score_value: 得分值（来自 Excel 单元格）
            row_idx: 行索引（用于错误提示）
            col_idx: 列索引（用于错误提示）

        Returns:
            解析后的得分（数值或区间数）
        """
        # 转换为字符串
        if isinstance(score_value, (int, float)):
            return float(score_value)

        score_str = str(score_value).strip()

        # 尝试解析为区间数
        if ',' in score_str:
            parts = score_str.split(',')
            if len(parts) != 2:
                raise ValueError(f"区间数格式错误，应为 'a,b' 或 '[a,b]'")

            lower_str = parts[0].strip().strip('[]').strip()
            upper_str = parts[1].strip().strip('[]').strip()

            try:
                lower = float(lower_str)
                upper = float(upper_str)
            except ValueError as e:
                raise ValueError(
                    f"区间数值无效：'{lower_str}', '{upper_str}'"
                ) from e

            # 导入 Interval 类
            from ..interval import Interval
            return Interval(lower, upper)

        # 尝试解析为单个数值
        try:
            return float(score_str)
        except ValueError as e:
            raise ValueError(
                f"无法解析得分值 '{score_str}'，"
                f"支持格式：数值（如 85）或区间数（如 80,90 或 [80,90]）"
            ) from e

    def validate(self, data: dict[str, Any]) -> bool:
        """验证 Excel 配置数据

        Args:
            data: 配置数据字典

        Returns:
            True（基本验证通过）
        """
        # 基本验证
        if not isinstance(data, dict):
            return False

        required_keys = ['alternatives', 'criteria', 'matrix']
        return all(key in data for key in required_keys)
